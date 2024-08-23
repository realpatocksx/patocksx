from flask import Flask, render_template, request, redirect
import openpyxl
import logging, os
from datetime import datetime

app = Flask(__name__, template_folder=os.path.abspath('DuckChat/client'), static_folder=os.path.abspath('DuckChat/client'))

# Configurar o arquivo de log
logging.basicConfig(filename='error.log', level=logging.ERROR, 
                    format='%(asctime)s:%(levelname)s:%(message)s')

# Função para salvar dados no arquivo Excel
def save_to_excel(username, message):
    file_path = 'chat_data.xlsx'

    try:
        # Carregar o arquivo Excel ou criar um novo
        try:
            workbook = openpyxl.load_workbook(file_path)
            sheet = workbook.active
        except FileNotFoundError:
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            # Criar cabeçalhos
            sheet.append(['Username', 'Message', 'Timestamp'])

        # Adicionar nova linha com o usuário, mensagem e horário
        sheet.append([username, message, datetime.now().strftime('%Y-%m-%d %H:%M:%S')])

        # Salvar o arquivo Excel
        workbook.save(file_path)
    except Exception as e:
        logging.error(f"Erro ao salvar mensagem no Excel: {e}")

# Função para ler dados do arquivo Excel
def read_from_excel():
    file_path = 'chat_data.xlsx'
    data = []

    try:
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active

        # Ler os dados do Excel e armazenar em uma lista
        for row in sheet.iter_rows(min_row=2, values_only=True):  # Ignora o cabeçalho
            data.append({
                'username': row[0],
                'message': row[1],
                'timestamp': row[2]
            })
    except FileNotFoundError:
        # Se o arquivo não existir, retorna uma lista vazia
        pass
    except Exception as e:
        logging.error(f"Erro ao ler mensagens do Excel: {e}")

    return data

@app.route('/')
def index():
    # Carregar mensagens do arquivo Excel
    messages = read_from_excel()
    return render_template('chat.html', messages=messages)

@app.route('/send_message', methods=['POST'])
def send_message():
    username = request.form.get('username')
    message = request.form.get('message')

    if username and message:
        # Salvar a mensagem no Excel
        save_to_excel(username, message)

    return redirect('/')

if __name__ == '__main__':
    try:
        app.run(debug=True)
    except Exception as e:
        logging.error(f"Erro ao iniciar o aplicativo Flask: {e}")

'''


'''